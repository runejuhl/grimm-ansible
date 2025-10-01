from __future__ import annotations
from typing import Dict
from collections import defaultdict

DOCUMENTATION = """
    name: petardo.ess.firewall
    version_added: "2.4"
    short_description: Load firewall rules from enterprise spreadsheets
    requirements:
        - Enabled in configuration
    description:
        - Loads spreadsheet rows into corresponding groups/hosts in group_vars/ and host_vars/ directories.
    options:
      stage:
        ini:
          - key: stage
            section: vars_petardo_ess_firewall
        env:
          - name: ANSIBLE_VARS_PLUGIN_STAGE
      _valid_extensions:
        default: [".xlsx", ".xls"]
        description:
          - "Check all of these extensions when looking for 'variable' files which should be YAML or JSON or vaulted versions of these."
          - 'This affects vars_files, include_vars, inventory and vars plugins among others.'
        env:
          - name: ANSIBLE_ESS_FILENAME_EXT
        ini:
          - key: ess_valid_extensions
            section: defaults
        type: list
        elements: string
    extends_documentation_fragment:
      - vars_plugin_staging
"""

import os
from ansible.errors import AnsibleParserError
from ansible.module_utils.common.text.converters import to_native
from ansible.plugins.vars import BaseVarsPlugin
from ansible.utils.path import basedir
from ansible.inventory.group import InventoryObjectType
from ansible.utils.vars import merge_hash
from ansible.parsing.dataloader import DataLoader

from openpyxl import load_workbook

recursivedict = lambda: defaultdict(recursivedict)

CANONICAL_PATHS = {}  # type: dict[str, str]
FOUND = {}  # type: dict[str, list[str]]
NAK = set()  # type: set[str]

default_group = "NO_GROUP"

FIREWALL_FIELD_MAP = {
    "Host": "host",
    "Source": "src",
    "Source port": "sport",
    "Destination": "dst",
    "Destination port": "dport",
    "Protocol": "proto",
    "Action": "action",
    "Comment": "comment",
}


class VarsModule(BaseVarsPlugin):
    is_stateless = True

    def _get_header(self, row) -> list[str]:
        return [x.value for x in row if x.value]

    def _get_dimensions(self, sheet) -> tuple[int, int]:
        """Get dimensions of sheet `sheet`.

        The openpyxl functions `Worksheet.max_column` and `Worksheet.max_row`
        only gives an approximate value; when a cell value is deleted, the sheet
        metadata isn't necessarily updated with the "real" dimensions.

        """
        maxcol = sheet.max_column
        maxrow = sheet.max_row

        # first row is a header, so we know it'll have a value
        col = maxcol
        for col in reversed(range(1, maxcol + 1)):
            if sheet.cell(1, col).value:
                maxcol = col + 1
                break

        for row in reversed(range(1, maxrow + 1)):
            for col in range(1, maxcol + 1):
                if sheet.cell(row, col).value:
                    return (row, maxcol)

        raise AnsibleParserError(f"unable to find dimensions of {sheet}")

    def load_file(self, file) -> Dict:
        data = {"firewall": {"rules": []}}

        wb = load_workbook(file)
        sheet = wb["Firewall"]  # if "sheet" in config else wb.active
        rows = list(sheet.rows)
        header = self._get_header(rows[0])

        if not header == [
            "Source",
            "Source port",
            "Destination",
            "Destination port",
            "Protocol",
            "Action",
            "Comment",
        ]:
            raise AnsibleParserError(f"invalid header in {file}: {header}")

        (maxrow, _maxcol) = self._get_dimensions(sheet)

        rules = data["firewall"]["rules"]
        for column in rows[1:maxrow]:

            row = {}

            for idx, var_name in enumerate(column):
                if column[idx].value is None:
                    continue

                match var_name.data_type:
                    case "s" | "n":
                        row[header[idx]] = var_name.value
                    case _:
                        raise AnsibleParserError(
                            f"invalid type '{var_name.data_type}' in field {var_name.coordinate} (value '{var_name.value}')"
                        )


            rules.append(row)

        return data

    def load_found_files(self, loader: DataLoader, data, found_files) -> Dict:
        for file in found_files:
            new_data = self.load_file(file)
            if new_data:
                data = merge_hash(data, new_data, list_merge="append_rp")

        return data

    def get_vars(self, loader, path, entities, cache=False) -> Dict:
        """FIXME: write this"""


        if not isinstance(entities, list):
            entities = [entities]

        # realpath is expensive
        try:
            realpath_basedir = CANONICAL_PATHS[path]
        except KeyError:
            CANONICAL_PATHS[path] = realpath_basedir = os.path.realpath(basedir(path))

        data = {}
        # data = recursivedict()
        for entity in entities:
            try:
                entity_name = entity.name
            except AttributeError:
                raise AnsibleParserError(
                    "Supplied entity must be Host or Group, got %s instead"
                    % (type(entity))
                )

            try:
                first_char = entity_name[0]
            except (TypeError, IndexError, KeyError):
                raise AnsibleParserError(
                    "Supplied entity must be Host or Group, got %s instead"
                    % (type(entity))
                )

            # avoid 'chroot' type inventory hostnames /path/to/chroot
            if first_char == os.path.sep:
                continue

            try:
                found_files = []
                # load vars
                try:
                    entity_type = entity.base_type
                except AttributeError:
                    raise AnsibleParserError(
                        "Supplied entity must be Host or Group, got %s instead"
                        % (type(entity))
                    )

                if entity_type is InventoryObjectType.HOST:
                    subdir = "host_vars"
                elif entity_type is InventoryObjectType.GROUP:
                    subdir = "group_vars"
                else:
                    raise AnsibleParserError(
                        "Supplied entity must be Host or Group, got %s instead"
                        % (type(entity))
                    )

                opath = os.path.join(realpath_basedir, subdir)
                key = "%s.%s" % (entity_name, opath)

                if os.path.isdir(opath):
                    FOUND[key] = found_files = loader.find_vars_files(
                        opath, entity_name, extensions=["", ".xlsx"]
                    )
                elif not os.path.exists(opath):
                    # cache missing dirs so we don't have to keep looking for things beneath the
                    NAK.add(opath)
                else:
                        "file",
                        "Found %s that is not a directory, skipping: %s"
                        % (subdir, opath),
                    )
                    # cache non-directory matches
                    NAK.add(opath)

                data = self.load_found_files(loader, data, found_files)

            except Exception as e:
                raise AnsibleParserError(to_native(e))

        return data
